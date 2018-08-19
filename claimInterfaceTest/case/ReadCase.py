from distutils.log import Log

import openpyxl
from src.InterfaceTest import *
from logs.Log import *

class ReadCase:
    def getCase(self,path1,path2):
        log = Log('F:\\test.log')
        try:
            #打开excel文件
            openFile = openpyxl.load_workbook(path1)
            log.info("成功打开测试用例")
        except BaseException as e:
            log.info("打开失败:",str(e))
        #打开指定sheet
        # sheet = openFile.sheetnames("TestCase")

        sheet = openFile.get_sheet_by_name("TestCase")
        print("获取指定工作表:",sheet.title)

        #遍历excel
        for i in range(2,sheet.max_row + 1):
            if sheet.cell(row = i, column = 10).value != 'Yes':
                continue
            requestData1 = sheet.cell(row = i, column = 1).value

            requestData2 = sheet.cell(row=i, column=2).value

            requestData3 = sheet.cell(row=i, column=3).value

            requestData4 = sheet.cell(row=i, column=4).value

            requestData5 = sheet.cell(row=i, column=5).value

            requestData6 = sheet.cell(row=i, column=6).value

            requestData7 = sheet.cell(row=i, column=7).value
            requestData7 = eval(requestData7)

            requestData8 = sheet.cell(row=i, column=8).value

            print(type(requestData1), requestData1)
            # requestData9 = sheet.cell(row=i, column=9).value
            # print(type(requestData9), requestData9)

            # headers = {"Content-Type":"application/x-www-form-urlencoded"}
            it = InterfaceTest()
            it.testRequest(requestData4,requestData7,requestData5,requestData6,
                           requestData8,i,sheet,requestData1,requestData2,log)

            openFile.save(path2)

if __name__ == '__main__':
    path1 = "F:\\testcase.xlsx"
    path2 = "F:\\testreport.xlsx"
    readcase1 = ReadCase()
    readcase1.getCase(path1, path2)